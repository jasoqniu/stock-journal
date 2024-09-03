import openpyxl
from openpyxl import Workbook
from openpyxl import workbook
from openpyxl import load_workbook
import yfinance as yf
import datetime 
from datetime import datetime as dtm
import investpy
import pandas as pd
from datetime import timedelta
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl.styles import Font, Color

#path
path = "C:\\Users\\Jason\\Downloads\\puefolder\\"

#set excel workbook
wb_AL = openpyxl.Workbook()
wb_MZ = openpyxl.Workbook()
wb_AL.remove(wb_AL["Sheet"])
wb_MZ.remove(wb_MZ["Sheet"])


#set date & time
current_datetime = dtm.now()
time_delta = timedelta(days=162)
past = current_datetime - time_delta
current_date = current_datetime.strftime("%Y-%m-%d")
date = past.strftime("%d-%m-%Y")


#stocklist.xlsx = https://www.idx.co.id/id/data-pasar/data-saham/daftar-saham/
excel_file = path + 'stocklist.xlsx'
df = pd.read_excel(excel_file)
stock_list = df['Kode'].tolist()


#input percentage 
percent_input = int(input("INPUT REQUESTED PERCENTAGE : "))
percent_input = percent_input/100
print(percent_input)

#looping stock list
for s in stock_list:
    kode = s
    first_letter = s[0]
    print("Symbol is : ",s)
    data = yf.download(kode + ".JK", start=past,end=current_datetime)
    data = data.sort_index(ascending=False)
    
    #filter volume under 1 000 000 000
    avg_volume_100 = data['Volume'].mean()
    avg_close_100 = data['Close'].mean()

    if avg_volume_100 * avg_close_100 < 1000000000:
        print(kode,"removed, volume under 1 000 000 000")
        continue
    
    #filter into A-L and M-Z
    if 'A' <= first_letter <= 'L':
        current_workbook = wb_AL
    elif 'M' <= first_letter <= 'Z':
        current_workbook = wb_MZ
    
    
    #create new sheet
    new_sheet = current_workbook.create_sheet(kode)
    sheet = current_workbook[kode]
    
    #data length
    data_length = len(data)
    print("AMOUNT OF DAYS = ",data_length)
    
    #set label for row=1
    sheet.cell(row=1, column=1).value = "Tanggal"
    sheet.cell(row=1, column=2).value = "Open"
    sheet.cell(row=1, column=3).value = "High"
    sheet.cell(row=1, column=4).value = "Low"
    sheet.cell(row=1, column=5).value = "Close"
    sheet.cell(row=1, column=6).value = "Volume"
    sheet.cell(row=1, column=8).value = "CH"
    sheet.cell(row=1, column=9).value = "CL"
    sheet.cell(row=1, column=10).value = "CC"
    sheet.cell(row=1, column=11).value = "Avg Harian"
    sheet.cell(row=1, column=12).value = "MA5"
    sheet.cell(row=1, column=13).value = "op=low"
    sheet.cell(row=1, column=14).value = "op=high"
    sheet.cell(row=1, column=15).value = "Prank"
    sheet.cell(row=1, column=16).value = "JJS OpLo"
    sheet.cell(row=1, column=17).value = "WR JJSOL"
    sheet.cell(row=1, column=18).value = "Prank%"
    sheet.cell(row=1, column=19).value = "OpLo9"
    sheet.cell(row=1, column=20).value = "WR OpLo9"
    
    
    #set width for row=1
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 13

    #color row=1
    for c in range(1,21):
        sheet.cell(row=1,column=c).fill = PatternFill(start_color="94c685",
                                        end_color="94c685",
                                        fill_type="solid")
        sheet.cell(row=1,column=c).font = Font(color="FFFFFF", bold=True, italic=False)
    
    #OHLCV
    rowvalue = 2
    for index,row in data.iterrows():
        open_value = row['Open']
        high_value = row['High']
        low_value = row['Low']
        close_value = row['Close']
        volume_value = row['Volume']
        tanggal_value = index.date()
        '''
        open_position = sheet.cell(row=rowvalue, column=2)
        high_position = sheet.cell(row=rowvalue, column=3) 
        low_position = sheet.cell(row=rowvalue, column=4) 
        close_position = sheet.cell(row=rowvalue, column=5)
        volume_position = sheet.cell(row=rowvalue, column=6)
        '''
        sheet.cell(row=rowvalue, column=1).value = tanggal_value
        sheet.cell(row=rowvalue, column=2).value = open_value
        sheet.cell(row=rowvalue, column=3).value = high_value
        sheet.cell(row=rowvalue, column=4).value = low_value
        sheet.cell(row=rowvalue, column=5).value = close_value
        sheet.cell(row=rowvalue, column=6).value = volume_value
        rowvalue += 1
    
    #Color OHLCV
    for rowvalue in range(2, (data_length+2)):
        if rowvalue % 2 == 0:
            for colvalue in range(1,7):
                cell = sheet.cell(row=rowvalue,column=colvalue)
                cell.fill = PatternFill(start_color="ffffff",
                                        end_color="ffffff",
                                        fill_type="solid")
        else:
            for colvalue in range(1,7):
                cell = sheet.cell(row=rowvalue,column=colvalue)
                cell.fill = PatternFill(start_color="a2c997",
                                        end_color="a2c997",
                                        fill_type="solid")        
    
    #CH
    rowvalue = 2
    for index,row in data.iterrows():
        high_today = row['High']
        close_yesterday = sheet.cell(row=rowvalue + 1, column=5).value
        if close_yesterday:
            ch_value = (high_today - close_yesterday) / close_yesterday
            sheet.cell(row=rowvalue, column=8).value = ch_value
            #data.at[index, 'CH'] = ch_value
        sheet.cell(row=rowvalue, column=8).number_format = numbers.FORMAT_PERCENTAGE_00
        rowvalue += 1
    
    #CL
    rowvalue = 2
    for index, row in data.iterrows():
        low_today = row['Low']
        close_yesterday = sheet.cell(row=rowvalue + 1, column=5).value
        if close_yesterday:
            cl_value = (low_today-close_yesterday) / close_yesterday
            sheet.cell(row=rowvalue, column=9).value = cl_value
        sheet.cell(row=rowvalue, column=9).number_format = numbers.FORMAT_PERCENTAGE_00
        rowvalue += 1

    #CC
    rowvalue = 2
    for index, row in data.iterrows():
        close_yesterday = sheet.cell(row=rowvalue + 1, column=5).value
        close_today = sheet.cell(row=rowvalue, column=5).value
        if close_yesterday:
            cc_value = (close_today - close_yesterday) / close_yesterday
            sheet.cell(row=rowvalue, column=10).value = cc_value
        sheet.cell(row=rowvalue, column=10).number_format = numbers.FORMAT_PERCENTAGE_00
        rowvalue += 1
    
    #Color CH CL CC
    rowvalue = 2
    for r in range(data_length):
        ch_value = sheet.cell(row=rowvalue, column=8).value 
        cl_value = sheet.cell(row=rowvalue, column=9).value 
        cc_value = sheet.cell(row=rowvalue, column=10).value 

        if ch_value is not None and ch_value >= percent_input:
            sheet.cell(row=rowvalue,column=8).fill = PatternFill(start_color="6fe26f",
                                                                 end_color="6fe26f",
                                                                   fill_type="solid")
        if cl_value is not None and cl_value <= -percent_input:
            sheet.cell(row=rowvalue,column=9).fill = PatternFill(start_color="e2746f",
                                                                 end_color="e2746f",
                                                                   fill_type="solid")
            sheet.cell(row=rowvalue, column=9).font = Font(color="FFFFFF", bold=True, italic=False)
            
        if cc_value is not None and cc_value <= -percent_input:
             sheet.cell(row=rowvalue,column=10).fill = PatternFill(start_color="e2746f",
                                                                 end_color="e2746f",
                                                                   fill_type="solid")
             sheet.cell(row=rowvalue, column=10).font = Font(color="FFFFFF", bold=True, italic=False)
        rowvalue += 1
        
    
    #Avg Harian
    rowvalue = 2
    for index,row in data.iterrows():
        open_value = row['Open']
        high_value = row['High']
        low_value = row['Low']
        close_value = row['Close']
        avg_harian = (open_value + high_value + low_value + close_value) / 4
        sheet.cell(row=rowvalue,column=11).value = avg_harian
        rowvalue += 1
    
    #MA5
    for r in range(2,(data_length+2)):
        if r <= (data_length-4):
            sum_ma5 = 0
            for i in range(5):
                sum_ma5 += sheet.cell(row=r+i,column=11).value
            ma5_value = sum_ma5 / 5
            sheet.cell(row=r,column=12).value = ma5_value
    
    #op=lo op=high prank
    prank_true = 0
    prank_false = 0
    for r in range(2,(data_length+2)):
        open_value = sheet.cell(row=r,column=2).value
        high_vaue = sheet.cell(row=r,column=3).value
        low_value = sheet.cell(row=r,column=4).value
        ch_value = sheet.cell(row=r, column=8).value 

        
        #op==lo
        if sheet.cell(row=r,column=2).value == sheet.cell(row=r,column=4).value:
            sheet.cell(row=r,column=13).value = "YES"
        else:
            sheet.cell(row=r,column=13).value = "---"
            
        #op==high
        if  sheet.cell(row=r,column=2).value == sheet.cell(row=r,column=3).value:
            sheet.cell(row=r,column=14).value = "YES"
            
            #prank
            if sheet.cell(row=r, column=8).value is not None and sheet.cell(row=r, column=8).value >= percent_input:
                sheet.cell(row=r,column=15).value = "PRANK"
                prank_true += 1
                #print("PRANK = TRUE")
            else:
                sheet.cell(row=r,column=15).value = "---"
                prank_false +=1
            
        else:
            sheet.cell(row=r,column=14).value = "---"
        
    #prank%
    prank_total = prank_true + prank_false
    if prank_true == 0:
        prank_percent = 0
    else:
        prank_percent = prank_true / prank_total
    #print("PRANK% = ",prank_percent)
    sheet.cell(row=2, column=18).value  = prank_percent
    sheet.cell(row=2, column=18).number_format = numbers.FORMAT_PERCENTAGE_00       
        
    #JJSOplo 
    jjsoplo_wins = 0
    jjsoplo_count = 0
    for r in range(2,(data_length+2)):
        open_today = sheet.cell(row=r,column=2).value
        low_today = sheet.cell(row=r,column=4).value  
        ch_tomorrow = sheet.cell(row=r-1, column=8).value 
        
        if open_today == low_today and r != 2:
            jjsoplo_count +=1
            if ch_tomorrow is not None and ch_tomorrow >= percent_input:
                sheet.cell(row=r,column=16).value = "WIN"
                jjsoplo_wins += 1
            else:
                sheet.cell(row=r,column=16).value = "LOOSE"
        else:
            sheet.cell(row=r,column=16).value = "---"
    #JJSOplo WR%        
    if jjsoplo_wins == 0:
        jjsoplo_wr = 0
    else:
        jjsoplo_wr = jjsoplo_wins / jjsoplo_count
    sheet.cell(row=2,column=17).value = jjsoplo_wr
    sheet.cell(row=2,column=17).number_format = numbers.FORMAT_PERCENTAGE_00

    #OpLo9
    oplo9_win = 0
    oplo9_count = 0
    for r in range(2,(data_length+2)):
        open_value = sheet.cell(row=r,column=2).value
        high_value = sheet.cell(row=r,column=3).value
        low_value = sheet.cell(row=r,column=4).value
        
        if open_value == low_value:
            oplo9_count +=1
            if (high_value / open_value) >= 1 + (percent_input):
                sheet.cell(row=r,column=19).value = "WIN"
                oplo9_win +=1
            else:
                sheet.cell(row=r,column=19).value = "LOOSE"
        else:
            sheet.cell(row=r,column=19).value = "---"
    
    #WR OpLo9
    if oplo9_win == 0:
        oplo9_wr = 0
    else:
        oplo9_wr = oplo9_win/oplo9_count
    sheet.cell(row=2,column=20).value = oplo9_wr
    sheet.cell(row=2,column=20).number_format = numbers.FORMAT_PERCENTAGE_00
    
#save to A-L & M-Z workbook
wb_AL.save(path + "new_AL_" + str(percent_input) + ".xlsx")
wb_MZ.save(path + "new_MZ_" + str(percent_input) + ".xlsx")



wb_results = openpyxl.Workbook()
wb_results.remove(wb_results["Sheet"]) 


    
#----------FILTERING----------#
#///FILTER JJSOL///
def filter_wrjjsol(wb, arr):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        vol_sum = 0
        for r in range(2, 102):
            if sheet.cell(row=r, column=6).value is not None:
                vol_sum += sheet.cell(row=r, column=6).value
        vol_avg = vol_sum / 100
        
        if vol_avg < 50000:
            continue
        
        wrjjsol = sheet.cell(row=2, column=17).value

        if wrjjsol > 0.6:
            arr.append((sheet_name, wrjjsol))
# arr for filtered stocks
arr_filter_wrrjjsol = []
# call function
filter_wrjjsol(wb_AL, arr_filter_wrrjjsol)
filter_wrjjsol(wb_MZ, arr_filter_wrrjjsol)

# sort stock
arr_filter_wrrjjsol.sort(key=lambda x: x[1], reverse=True)


#save into results workbook
new_sheet = wb_results.create_sheet("WRJJSOL")
new_sheet.append(["Kode", "WR% JJSOL"])

rowvalue = 2
for kode, wr in arr_filter_wrrjjsol:
    new_sheet.append([kode, wr])
    new_sheet.cell(row=rowvalue, column=2).number_format = numbers.FORMAT_PERCENTAGE_00
    rowvalue+= 1



#///FILTER OPLO9///#
def fiter_wroplo9(wb, arr):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        vol_sum = 0
        for r in range(2, 102):
            if sheet.cell(row=r, column=6).value is not None:
                vol_sum += sheet.cell(row=r, column=6).value
        vol_avg = vol_sum / 100
        
        if vol_avg < 50000:
            continue
        
        wroplo9 = sheet.cell(row=2, column=20).value
        if wroplo9 > 0.6:
            arr.append((sheet_name, wroplo9))
# arr for filtered stocks
arr_filter_wroplo9 = []
# call function
fiter_wroplo9(wb_AL, arr_filter_wroplo9)
fiter_wroplo9(wb_MZ, arr_filter_wroplo9)

# sort stock
arr_filter_wroplo9.sort(key=lambda x: x[1], reverse=True)

new_sheet = wb_results.create_sheet("WROPLO9")
new_sheet.append(["Kode", "WR% OpLo9"])


rowvalue = 2
for kode, wr in arr_filter_wroplo9:
    new_sheet.append([kode, wr])
    new_sheet.cell(row=rowvalue, column=2).number_format = numbers.FORMAT_PERCENTAGE_00
    rowvalue+= 1
#-----------------------------#

#--------------MARUBOZU---------------#
new_sheet = wb_results.create_sheet("MRBZ FULL BULLISH")
new_sheet.cell(row=1,column=1).value = "Kode"
def fullbullish(wb,output_sheet):
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        
        val_arr = []
        for r in range(2,102):
            close = sheet.cell(row=r,column=5).value
            vol = sheet.cell(row=r,column=6).value
            if vol is not None and close is not None:
                val = vol + close
                val_arr.append(val)
        avg_val = sum(val_arr) / len(val_arr)
        
        open = sheet.cell(row=2,column=2).value
        high = sheet.cell(row=2,column=3).value
        low = sheet.cell(row=2,column=4).value
        close = sheet.cell(row=2,column=5).value
        
        if open == low and close == high and avg_val > 10000000:
            output_sheet.append([sheetname])

fullbullish(wb_AL,new_sheet)
fullbullish(wb_MZ,new_sheet)

new_sheet = wb_results.create_sheet("MRBZ CLOSE BULLISH")
new_sheet.cell(row=1,column=1).value = "Kode"
def closebullish(wb,output_sheet):
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        
        val_arr = []
        for r in range(2,102):
            close = sheet.cell(row=r,column=5).value
            vol = sheet.cell(row=r,column=6).value
            if vol is not None and close is not None:
                val = vol + close
                val_arr.append(val)
        avg_val = sum(val_arr) / len(val_arr)
        
        open = sheet.cell(row=2,column=2).value
        high = sheet.cell(row=2,column=3).value
        close = sheet.cell(row=2,column=5).value
        
        if close == high and open < close and avg_val > 1000000000:
            output_sheet.append([sheetname])
            
closebullish(wb_AL,new_sheet)
closebullish(wb_MZ,new_sheet)

#-------------------------------------#

#------------BOW--------------#
def bow(wb):
    bow_data = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        total_red = 0
        bow_count = 0
        
        for r in range(2, 102):
            cl_value = sheet.cell(row=r, column=9).value
            cc_value = sheet.cell(row=r, column=10).value
            if cl_value is not None and cc_value is not None:
                if cl_value < -percent_input:  
                    total_red += 1
                    if cl_value < cc_value: 
                        bow_count += 1
        
        if total_red > 0:
            bow_rate = bow_count / total_red
            bow_data.append([sheet_name, bow_count, bow_rate])
            
    return bow_data

new_sheet = wb_results.create_sheet("BOW")
new_sheet.append(["Kode", "BOWs", "BOW%"])

bow_data_al = bow(wb_AL)
bow_data_mz= bow(wb_MZ)

bow_data = bow_data_al + bow_data_mz
bow_data.sort(key=lambda x: x[2], reverse=True)

for i, row in enumerate(bow_data,2):
    new_sheet.append(row)
    new_sheet.cell(row=i, column=3).number_format = numbers.FORMAT_PERCENTAGE_00

#-----------------------------#

#-------------POLA----------------#
def pola(wb, arr1):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        pola1 = 0
        pola2 = 0
        pola3 = 0
        pola4 = 0
        pola5 = 0
        streakhijau = 0
        arr_hijau = []
        for r in range(2, 102):
            ch_today = sheet.cell(row=r, column=8).value 
            ch_yesterday = sheet.cell(row=r + 1, column=8).value 
            #CARI POLA
            if ch_today is not None and ch_yesterday is not None:
                if ch_today >= percent_input and ch_yesterday < percent_input:  # Hijau putih
                    if sheet.cell(row=r + 2, column=8).value is not None and sheet.cell(row=r + 2, column=8).value >= percent_input:  # Hijau putih hijau
                        pola1 += 1
                    elif sheet.cell(row=r + 2, column=8).value is not None and sheet.cell(row=r + 2, column=8).value < percent_input:  # Hijau putih2
                        if sheet.cell(row=r + 3, column=8).value is not None and sheet.cell(row=r + 3, column=8).value >= percent_input:  # Hijau putih2 hijau
                            pola2 += 1
                        elif sheet.cell(row=r + 3, column=8).value is not None and sheet.cell(row=r + 3, column=8).value < percent_input:  # Hijau putih3
                            if sheet.cell(row=r + 4, column=8).value is not None and sheet.cell(row=r + 4, column=8).value >= percent_input:  # Hijau putih3 hijau
                                pola3 += 1
                            elif sheet.cell(row=r + 4, column=8).value is not None and sheet.cell(row=r + 4, column=8).value < percent_input:  # Hijau putih4 
                                if sheet.cell(row=r + 5, column=8).value is not None and sheet.cell(row=r + 5, column=8).value >= percent_input:  # Hijau putih4 hijau
                                    pola4 += 1
                                elif sheet.cell(row=r + 5, column=8).value is not None and sheet.cell(row=r + 5, column=8).value < percent_input:  # Hijau putih5
                                    if sheet.cell(row=r + 6, column=8).value is not None and sheet.cell(row=r + 6, column=8).value >= percent_input:  # Hijau putih5 hijau
                                        pola5 += 1
            #PARADE HIJAU
            if ch_today is not None and ch_yesterday is not None and ch_today >= percent_input and ch_yesterday >= percent_input:
                streakhijau += 1
            elif ch_today is not None and ch_yesterday is not None and ch_today >= percent_input and ch_yesterday < percent_input:
                streakhijau +=1
                arr_hijau.append(streakhijau)
                streakhijau = 0
            elif r == 101:
                streakhijau +=1
                arr_hijau.append(streakhijau)
                streakhijau = 0
        print(sheet, arr_hijau)
        sumhijau = sum(arr_hijau)
        lenhijau = len(arr_hijau)
        paradehijau = sumhijau/lenhijau
        if pola4 > 2 or pola5 > 0:
            pola_type = "---"
        elif pola1 > 2*pola2:
            pola_type = "Pola 1"
        elif pola2 > 2*pola3:
            pola_type = "Pola 2"
        elif pola3 > 2*pola4:
            pola_type = "Pola 3"
        else:
            pola_type = "---"
        arr1.append((sheet_name,pola_type,paradehijau,pola1,pola2,pola3,pola4,pola5))
        
arr_pola = []
pola(wb_AL,arr_pola)
pola(wb_MZ,arr_pola)

urut_pola = {"Pola 1": 1, "Pola 2": 2, "Pola 3": 3, "---": 4}
arr_pola.sort(key=lambda x: urut_pola[x[1]])


new_sheet = wb_results.create_sheet("POLA")
new_sheet.append(["Kode","Tipe Pola","Parade Hijau", "Pola1","Pola2","Pola3","Pola4","Pola5"])
rowvalue = 2
for kode,p,ph,p1,p2,p3,p4,p5 in arr_pola:
    new_sheet.append([kode,p,ph,p1,p2,p3,p4,p5])
    new_sheet.cell(row=rowvalue, column=3).number_format = '0.00'
    rowvalue+= 1
#---------------------------------#
wb_results.save(path + "RESULTS_" + str(percent_input) + ".xlsx")
